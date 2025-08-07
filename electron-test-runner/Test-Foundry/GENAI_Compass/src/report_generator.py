import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, List
import os
from datetime import datetime
from pathlib import Path

class ReportGenerator:
    """
    FINAL WORKING: Report generator with all sheets restored and proper percentage conversion
    """
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.color_scheme = config.get('reporting', {}).get('color_coding', {
            'good': '92D050',
            'acceptable': 'FFC000', 
            'poor': 'FF0000'
        })
        
    def generate_report(self, results_df: pd.DataFrame, summary_stats: Dict[str, Any], 
                       metadata: Dict[str, Any], input_filename: str) -> None:
        """
        FINAL WORKING: Generate report with proper numeric to percentage conversion
        """
        print("Generating comprehensive report...")
        
        # Create output directory and filename
        output_dir = Path("C:/GENAI_Compass/Reports")
        output_dir.mkdir(parents=True, exist_ok=True)
        
        input_name = Path(input_filename).stem
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"{input_name}_EvaluationReport_{timestamp}.xlsx"
        output_path = output_dir / output_filename
        
        # Store original values and input filename
        self.original_results_df = results_df.copy()
        self.input_filename = input_name
        
        # Log original overall scores (should be numeric 0.0 to 1.0)
        print("FINAL WORKING - Original overall scores (should be numeric):")
        for i, score in enumerate(results_df['overall_score']):
            print(f"  Row {i}: {score} (type: {type(score)})")
        
        # Convert to percentages ONLY for display
        enhanced_results_df = self._enhance_results_dataframe_working(results_df)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # Create all sheets (restored)
        self._create_results_sheet(wb, enhanced_results_df, metadata)
        self._create_summary_sheet(wb, summary_stats, metadata)
        self._create_metrics_info_sheet(wb, metadata)
        self._create_dependencies_sheet(wb, metadata)
        self._create_interpretation_sheet(wb, metadata)
        
        # Save workbook
        wb.save(str(output_path))
        print(f"✓ Report saved to {output_path}")
    
    def _enhance_results_dataframe_working(self, results_df: pd.DataFrame) -> pd.DataFrame:
        """
        FINAL WORKING: Convert NUMERIC scores (0.0-1.0) to percentages for display
        """
        enhanced_df = results_df.copy()
        
        print("FINAL WORKING: Converting numeric scores to percentages...")
        
        # Process score columns (excluding remarks)
        score_columns = [col for col in enhanced_df.columns if '_score' in col and '_remarks' not in col]
        
        for col in score_columns:
            print(f"Processing column: {col}")
            
            if 'toxicity' in col.lower():
                # Handle toxicity (convert to safety percentage)
                enhanced_df[col] = enhanced_df[col].apply(
                    lambda x: f"{(1 - float(x)) * 100:.2f}%" if pd.notna(x) and isinstance(x, (int, float)) else "N/A"
                )
            else:
                # Handle normal metrics (convert to percentage)
                enhanced_df[col] = enhanced_df[col].apply(
                    lambda x: f"{float(x) * 100:.2f}%" if pd.notna(x) and isinstance(x, (int, float)) else "N/A"
                )
        
        # Handle overall_score (should be numeric between 0.0 and 1.0)
        if 'overall_score' in enhanced_df.columns:
            print("FINAL WORKING: Converting overall_score column...")
            
            def convert_overall_to_percentage(value):
                print(f"  Converting: {value} (type: {type(value)})")
                
                # Handle numeric values (expected case)
                if pd.notna(value) and isinstance(value, (int, float)):
                    if value > 0:  # Valid score
                        percentage = f"{float(value) * 100:.2f}%"
                        print(f"    Result: {percentage}")
                        return percentage
                    else:
                        print("    Value <= 0, returning N/A")
                        return "N/A"
                
                # Handle already converted percentages (shouldn't happen)
                elif isinstance(value, str) and '%' in value:
                    print(f"    Already percentage: {value}")
                    return value
                
                # Handle invalid values
                else:
                    print(f"    Invalid value, returning N/A")
                    return "N/A"
            
            enhanced_df['overall_score'] = enhanced_df['overall_score'].apply(convert_overall_to_percentage)
            
            # Log final conversions
            print("FINAL WORKING: Overall score after conversion:")
            for i, score in enumerate(enhanced_df['overall_score']):
                print(f"  Row {i}: {score}")
        
        return enhanced_df
    
    def _create_results_sheet(self, wb: Workbook, results_df: pd.DataFrame, metadata: Dict[str, Any]) -> None:
        """Create detailed results sheet with custom header and formatting"""
        ws = wb.create_sheet("Detailed Results")
        
        # Add custom title with input filename
        ws['A1'] = f"{self.input_filename} - NLP Quality Metrics Evaluation Results"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        ws.merge_cells('A1:Z1')
        
        # Add timestamp
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        ws['A2'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        
        # Add data starting from row 4
        for r_idx, row in enumerate(dataframe_to_rows(results_df, index=False, header=True), 4):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply formatting to header row
                if r_idx == 4:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
                
                # Apply formatting to all data cells
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
                    
                    # Apply color coding to score columns using original values
                    if r_idx > 4 and self._is_score_column(results_df.columns[c_idx-1]):
                        original_value = self._get_original_score_value(r_idx - 5, c_idx - 1)
                        self._apply_score_color(cell, original_value, results_df.columns[c_idx-1], metadata)
        
        # Auto-adjust column widths
        for col_num in range(1, len(results_df.columns) + 1):
            column_letter = self._get_column_letter(col_num)
            
            # Special handling for remarks columns
            if '_remarks' in str(results_df.columns[col_num-1]):
                ws.column_dimensions[column_letter].width = 80
                # Set row height for remarks
                for row_num in range(5, len(results_df) + 5):
                    ws.row_dimensions[row_num].height = 120
            else:
                # Check all cells in this column
                max_length = 0
                for row_num in range(4, len(results_df) + 5):
                    try:
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 4, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders to data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws[f'A4:Z{4 + len(results_df)}']:
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
    
    def _create_summary_sheet(self, wb: Workbook, summary_stats: Dict[str, Any], metadata: Dict[str, Any]) -> None:
        """Create complete summary statistics sheet"""
        ws = wb.create_sheet("Summary Statistics")
        
        # Title
        ws['A1'] = f"{self.input_filename} - Evaluation Summary Statistics"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        ws.merge_cells('A1:E1')
        
        row = 3
        
        # Overall statistics
        ws[f'A{row}'] = "Overall Performance"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        row += 1
        
        overall_stats = summary_stats.get('overall_stats', {})
        stats_data = [
            ['Metric', 'Value'],
            ['Total Test Cases', summary_stats.get('total_cases', 0)],
            ['Average Score', f"{overall_stats.get('mean_score', 0) * 100:.2f}%"],
            ['Median Score', f"{overall_stats.get('median_score', 0) * 100:.2f}%"],
            ['Standard Deviation', f"{overall_stats.get('std_score', 0) * 100:.2f}%"],
            ['Minimum Score', f"{overall_stats.get('min_score', 0) * 100:.2f}%"],
            ['Maximum Score', f"{overall_stats.get('max_score', 0) * 100:.2f}%"]
        ]
        
        for data_row in stats_data:
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
                if data_row == stats_data[0]:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            row += 1
        
        row += 2
        
        # Grade distribution
        ws[f'A{row}'] = "Grade Distribution"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        row += 1
        
        grade_dist = summary_stats.get('grade_distribution', {})
        headers = ['Grade', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        row += 1
        
        total_cases = summary_stats.get('total_cases', 1)
        for grade in ['A+', 'A', 'B', 'C', 'D', 'F']:
            count = grade_dist.get(grade, 0)
            percentage = (count / total_cases) * 100 if total_cases > 0 else 0
            
            ws.cell(row=row, column=1, value=grade).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.cell(row=row, column=2, value=count).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            row += 1
        
        row += 2
        
        # Individual metric averages
        ws[f'A{row}'] = "Individual Metric Performance"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        row += 1
        
        headers = ['Metric', 'Average Score', 'Weight', 'Contribution', 'Performance Rating']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        row += 1
        
        metric_averages = summary_stats.get('metric_averages', {})
        for metric_name, avg_data in metric_averages.items():
            avg_score = avg_data.get('mean', 0)
            weight = metadata.get(metric_name, {}).get('weight', 0)
            contribution = avg_score * weight
            rating = self._get_performance_rating(avg_score, metric_name, metadata)
            
            ws.cell(row=row, column=1, value=metric_name.replace('_', ' ').title()).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            
            # Display score as percentage
            if 'toxicity' in metric_name.lower():
                score_cell = ws.cell(row=row, column=2, value=f"{(1 - avg_score) * 100:.2f}%")
            else:
                score_cell = ws.cell(row=row, column=2, value=f"{avg_score * 100:.2f}%")
            
            score_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            self._apply_score_color(score_cell, avg_score, f"{metric_name}_score", metadata)
            
            ws.cell(row=row, column=3, value=f"{weight * 100:.1f}%").alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.cell(row=row, column=4, value=f"{contribution * 100:.2f}%").alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.cell(row=row, column=5, value=rating).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            row += 1
        
        # Auto-adjust column widths
        for col_num in range(1, 6):
            column_letter = self._get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 25
    
    def _create_metrics_info_sheet(self, wb: Workbook, metadata: Dict[str, Any]) -> None:
        """Create complete metrics information sheet"""
        ws = wb.create_sheet("Metrics Information")
        
        # Title
        ws['A1'] = f"{self.input_filename} - Metrics Configuration and Details"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        ws.merge_cells('A1:G1')
        
        row = 3
        
        # Headers
        headers = ['Metric', 'Description', 'Weight', 'Range', 'Higher is Better', 'Enabled', 'Interpretation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        row += 1
        
        # Metric details
        for metric_name, meta in metadata.items():
            ws.cell(row=row, column=1, value=metric_name.replace('_', ' ').title()).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.cell(row=row, column=2, value=meta.get('description', '')).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.cell(row=row, column=3, value=f"{meta.get('weight', 0) * 100:.1f}%").alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            
            # Update range to show percentages
            if 'toxicity' in metric_name.lower():
                ws.cell(row=row, column=4, value="[0% - 100%] (Safety Score)").alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            else:
                ws.cell(row=row, column=4, value="[0% - 100%]").alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            
            ws.cell(row=row, column=5, value='Yes' if meta.get('higher_is_better', True) else 'No').alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.cell(row=row, column=6, value='Yes' if meta.get('enabled', False) else 'No').alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            
            # Interpretation
            interpretation = meta.get('interpretation', {})
            interp_text = '; '.join([f"{k}: {v}" for k, v in interpretation.items()])
            ws.cell(row=row, column=7, value=interp_text).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            
            row += 1
        
        # Auto-adjust column widths
        for col_num in range(1, 8):
            column_letter = self._get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 30
    
    def _create_dependencies_sheet(self, wb: Workbook, metadata: Dict[str, Any]) -> None:
        """Create complete dependencies information sheet"""
        ws = wb.create_sheet("Dependencies")
        
        # Title
        ws['A1'] = f"{self.input_filename} - Metric Dependencies"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        ws.merge_cells('A1:D1')
        
        # Description
        ws['A3'] = "This sheet shows the required input columns for each metric calculation."
        ws['A3'].font = Font(italic=True)
        ws['A3'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        ws.merge_cells('A3:D3')
        
        row = 5
        
        # Headers
        headers = ['Metric', 'Required Columns', 'Optional Columns', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        row += 1
        
        # Dependency information
        for metric_name, meta in metadata.items():
            dependencies = meta.get('dependencies', [])
            
            required_cols = []
            optional_cols = []
            notes = []
            
            for dep in dependencies:
                if dep == 'Context':
                    optional_cols.append(dep)
                    notes.append("Context improves accuracy but is optional")
                else:
                    required_cols.append(dep)
            
            if meta.get('reverse_scoring', False):
                notes.append("Displayed as Safety Score (100% - Toxicity%)")
            
            ws.cell(row=row, column=1, value=metric_name.replace('_', ' ').title()).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.cell(row=row, column=2, value=', '.join(required_cols)).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.cell(row=row, column=3, value=', '.join(optional_cols) if optional_cols else 'None').alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.cell(row=row, column=4, value='; '.join(notes) if notes else 'Standard metric').alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            
            row += 1
        
        # Add summary table
        row += 2
        ws[f'A{row}'] = "Column Mapping Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        row += 1
        
        column_mapping = [
            ['Excel Column', 'Purpose', 'Used By'],
            ['Prompt', 'Original question/instruction', 'Answer Relevance'],
            ['Reference_Response', 'Expected/correct answer', 'BLEU, ROUGE, METEOR, BERT Score, Accuracy'],
            ['Generated_Response', 'AI-generated response', 'All metrics'],
            ['Context', 'Additional background info', 'Answer Relevance (optional)'],
            ['Enable_Flag', 'Y/N to run test', 'Test execution control']
        ]
        
        for data_row in column_mapping:
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
                if data_row == column_mapping[0]:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            row += 1
        
        # Auto-adjust column widths
        for col_num in range(1, 4):
            column_letter = self._get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 30
    
    def _create_interpretation_sheet(self, wb: Workbook, metadata: Dict[str, Any]) -> None:
        """Create complete score interpretation guide sheet"""
        ws = wb.create_sheet("Score Interpretation")
        
        # Title
        ws['A1'] = f"{self.input_filename} - Score Interpretation Guide"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        ws.merge_cells('A1:F1')
        
        # Description
        ws['A3'] = "This guide explains how to interpret the percentage scores for each metric."
        ws['A3'].font = Font(italic=True)
        ws['A3'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        ws.merge_cells('A3:F3')
        
        row = 5
        
        for metric_name, meta in metadata.items():
            # Metric title
            ws[f'A{row}'] = f"{metric_name.replace('_', ' ').title()}"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Description
            description = meta.get('description', '')
            if 'toxicity' in metric_name.lower():
                description += " (Displayed as Safety Score: 100% = completely safe, 0% = highly toxic)"
            ws[f'A{row}'] = description
            ws[f'A{row}'].font = Font(italic=True)
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            # Score ranges
            interpretation = meta.get('interpretation', {})
            if interpretation:
                headers = ['Performance Level', 'Score Range', 'Color Coding', 'Meaning']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
                row += 1
                
                # Add interpretation rows with percentage ranges
                color_mapping = {
                    'excellent': ('good', 'Exceptional performance'),
                    'good': ('good', 'Above average performance'),
                    'acceptable': ('acceptable', 'Satisfactory performance'),
                    'poor': ('poor', 'Below expectations')
                }
                
                for level, score_range in interpretation.items():
                    color_key, meaning = color_mapping.get(level, ('acceptable', 'Standard performance'))
                    
                    ws.cell(row=row, column=1, value=level.title()).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
                    
                    # Convert score range to percentage
                    if isinstance(score_range, str) and '-' in score_range:
                        try:
                            parts = score_range.split('-')
                            start = float(parts[0]) * 100
                            end = float(parts[1]) * 100
                            percentage_range = f"{start:.0f}% - {end:.0f}%"
                        except:
                            percentage_range = score_range
                    else:
                        percentage_range = f"{float(score_range) * 100:.0f}%+" if score_range else "N/A"
                    
                    ws.cell(row=row, column=2, value=percentage_range).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
                    
                    # Color coding cell
                    color_cell = ws.cell(row=row, column=3, value=color_key.title())
                    color_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
                    color_cell.fill = PatternFill(
                        start_color=self.color_scheme.get(color_key, 'FFFFFF'),
                        end_color=self.color_scheme.get(color_key, 'FFFFFF'),
                        fill_type='solid'
                    )
                    
                    ws.cell(row=row, column=4, value=meaning).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
                    row += 1
            
            # Special notes
            notes = []
            if meta.get('reverse_scoring', False):
                notes.append("⚠ Displayed as Safety Score (higher percentage = safer content)")
            notes.append("📊 All scores displayed as percentages for easier interpretation")
            
            if notes:
                row += 1
                ws[f'A{row}'] = "Notes: " + "; ".join(notes)
                ws[f'A{row}'].font = Font(italic=True, color='0066CC')
                ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
                ws.merge_cells(f'A{row}:F{row}')
            
            row += 3  # Add spacing between metrics
        
        # General guidelines
        ws[f'A{row}'] = "General Guidelines"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
        row += 1
        
        guidelines = [
            "• All scores are displayed as percentages (0% to 100%) for easier interpretation",
            "• Overall scores are calculated as weighted averages of individual metrics",
            "• Color coding provides visual indicators: Green (Good), Yellow (Acceptable), Red (Poor)",
            "• Focus on metrics with higher weights for overall performance improvement",
            "• Toxicity is shown as 'Safety Score' - higher percentages indicate safer content",
            "• Consider the detailed remarks for specific improvement suggestions"
        ]
        
        for guideline in guidelines:
            ws[f'A{row}'] = guideline
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)
            row += 1
        
        # Auto-adjust column widths
        for col_num in range(1, 7):
            column_letter = self._get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 30
    
    def _is_score_column(self, column_name: str) -> bool:
        """Check if column is a score column (not remarks)"""
        return ('_score' in str(column_name) or 'overall_score' in str(column_name)) and '_remarks' not in str(column_name)
    
    def _get_original_score_value(self, row_idx: int, col_idx: int) -> float:
        """Get original numeric score value for color coding"""
        try:
            if hasattr(self, 'original_results_df') and row_idx < len(self.original_results_df):
                value = self.original_results_df.iloc[row_idx, col_idx]
                if isinstance(value, (int, float)) and not pd.isna(value):
                    return float(value)
            return 0.5  # Neutral fallback for color coding
        except:
            return 0.5
    
    def _apply_score_color(self, cell, value: float, column_name: str, metadata: Dict[str, Any]) -> None:
        """Apply color coding based on score value"""
        try:
            metric_name = column_name.replace('_score', '').replace('overall', 'overall')
            
            if metric_name == 'overall':
                if value >= 0.8:
                    color = self.color_scheme['good']
                elif value >= 0.6:
                    color = self.color_scheme['acceptable']
                else:
                    color = self.color_scheme['poor']
            else:
                metric_meta = metadata.get(metric_name, {})
                thresholds = metric_meta.get('thresholds', {})
                reverse_scoring = metric_meta.get('reverse_scoring', False)
                
                if reverse_scoring:
                    if value <= thresholds.get('good', 0.2):
                        color = self.color_scheme['good']
                    elif value <= thresholds.get('acceptable', 0.5):
                        color = self.color_scheme['acceptable']
                    else:
                        color = self.color_scheme['poor']
                else:
                    if value >= thresholds.get('good', 0.7):
                        color = self.color_scheme['good']
                    elif value >= thresholds.get('acceptable', 0.5):
                        color = self.color_scheme['acceptable']
                    else:
                        color = self.color_scheme['poor']
            
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        except Exception:
            pass
    
    def _get_performance_rating(self, score: float, metric_name: str, metadata: Dict[str, Any]) -> str:
        """Get performance rating text based on score"""
        try:
            metric_meta = metadata.get(metric_name, {})
            thresholds = metric_meta.get('thresholds', {})
            reverse_scoring = metric_meta.get('reverse_scoring', False)
            
            if reverse_scoring:
                if score <= thresholds.get('good', 0.2):
                    return 'Excellent (Very Safe)'
                elif score <= thresholds.get('acceptable', 0.5):
                    return 'Good (Safe)'
                else:
                    return 'Needs Improvement (Potentially Unsafe)'
            else:
                if score >= thresholds.get('good', 0.7):
                    return 'Excellent'
                elif score >= thresholds.get('acceptable', 0.5):
                    return 'Good'
                else:
                    return 'Needs Improvement'
        except Exception:
            return 'Unknown'
    
    def _get_column_letter(self, col_num: int) -> str:
        """Convert column number to Excel column letter"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + col_num % 26) + result
            col_num //= 26
        return result