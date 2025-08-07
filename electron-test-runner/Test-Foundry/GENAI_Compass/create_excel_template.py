#!/usr/bin/env python3
"""
Excel Input Template Generator
Creates a comprehensive Excel template with examples and instructions
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def create_excel_template():
    """
    Create comprehensive Excel input template
    """
    print("Creating Excel input template...")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create instruction sheet
    create_instructions_sheet(wb)
    
    # Create template sheet
    create_template_sheet(wb)
    
    # Create examples sheet
    create_examples_sheet(wb)
    
    # Create column definitions sheet
    create_column_definitions_sheet(wb)
    
    # Save template
    template_path = "Excel_Input_Template.xlsx"
    wb.save(template_path)
    print(f"✓ Excel template created: {template_path}")
    
    return template_path

def create_instructions_sheet(wb):
    """Create instructions sheet"""
    ws = wb.create_sheet("📋 INSTRUCTIONS")
    
    # Title
    ws['A1'] = "NLP Evaluation Framework - Input Template Instructions"
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')
    
    # Instructions content
    instructions = [
        "",
        "OVERVIEW:",
        "This template helps you prepare data for the NLP Evaluation Framework.",
        "The framework evaluates AI-generated responses using 7 different metrics.",
        "",
        "REQUIRED SHEETS:",
        "• Use the 'Input Data' sheet for your actual test cases",
        "• See 'Examples' sheet for sample data formats",
        "• Check 'Column Definitions' for detailed field descriptions",
        "",
        "STEP-BY-STEP GUIDE:",
        "",
        "1. PREPARE YOUR DATA:",
        "   • Each row represents one test case to evaluate",
        "   • Fill in the required columns (marked with *)",
        "   • Optional columns can be left empty if not available",
        "",
        "2. REQUIRED COLUMNS (* = mandatory):",
        "   • Prompt* - The original question or instruction",
        "   • Reference_Response* - The expected/correct answer",
        "   • Generated_Response* - The AI-generated response to evaluate",
        "",
        "3. OPTIONAL COLUMNS:",
        "   • Context - Additional background information (improves Answer Relevance)",
        "   • Enable_Flag - Y/N to control which tests run (defaults to Y)",
        "",
        "4. DATA QUALITY TIPS:",
        "   • Keep responses clear and well-formatted",
        "   • Ensure reference responses are high-quality",
        "   • Include diverse test cases for comprehensive evaluation",
        "   • Add context when available for better relevance scoring",
        "",
        "5. RUNNING THE EVALUATION:",
        "   • Save this file as .xlsx format",
        "   • Use the 'Input Data' sheet as your input",
        "   • Run: python src/main.py --input your_file.xlsx",
        "",
        "6. WHAT YOU'LL GET:",
        "   • Comprehensive Excel report with multiple analysis sheets",
        "   • Individual metric scores (BLEU, ROUGE, METEOR, etc.)",
        "   • Color-coded performance indicators",
        "   • Detailed explanations and improvement suggestions",
        "",
        "METRICS EVALUATED:",
        "• BLEU Score - N-gram overlap similarity",
        "• ROUGE Score - Recall-oriented text evaluation", 
        "• METEOR Score - Considers synonyms and word order",
        "• BERT Score - Semantic similarity using AI embeddings",
        "• Accuracy - Factual correctness assessment",
        "• Answer Relevance - How well response addresses the prompt",
        "• Toxicity - Harmful content detection (lower is better)",
        "",
        "CONFIGURATION:",
        "• Metric weights and thresholds can be adjusted in config/metrics_config.yaml",
        "• Config file settings override Enable_Flag columns",
        "",
        "SUPPORT:",
        "• Check README.md for detailed documentation",
        "• Run 'python src/main.py create-sample' for quick test data",
        "• See example_usage.py for programmatic usage",
        "",
        f"Template created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ]
    
    # Add instructions
    for i, instruction in enumerate(instructions, 2):
        ws[f'A{i}'] = instruction
        if instruction.endswith(':') and instruction != "":
            ws[f'A{i}'].font = Font(bold=True, color='2F5496')
        elif instruction.startswith('•'):
            ws[f'A{i}'].font = Font(color='404040')
        elif instruction.startswith('   •'):
            ws[f'A{i}'].font = Font(color='606060')
    
    # Auto-adjust column width
    ws.column_dimensions['A'].width = 80

def create_template_sheet(wb):
    """Create main template sheet for data input"""
    ws = wb.create_sheet("📝 Input Data")
    
    # Title
    ws['A1'] = "NLP Evaluation Framework - Input Data Template"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:E1')
    
    # Instructions
    ws['A2'] = "Fill in your test cases below. Required columns are marked with *"
    ws['A2'].font = Font(italic=True, color='666666')
    ws.merge_cells('A2:E2')
    
    # Headers
    headers = [
        'Prompt*',
        'Reference_Response*', 
        'Generated_Response*',
        'Context',
        'Enable_Flag'
    ]
    
    # Header descriptions
    descriptions = [
        'The original question, instruction, or prompt',
        'Expected/correct answer or high-quality reference',
        'AI-generated response to be evaluated',
        'Optional: Additional context or background info',
        'Optional: Y/N to enable/disable this test case'
    ]
    
    # Add headers
    for col, (header, desc) in enumerate(zip(headers, descriptions), 1):
        # Header cell
        header_cell = ws.cell(row=4, column=col, value=header)
        header_cell.font = Font(bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Description cell
        desc_cell = ws.cell(row=5, column=col, value=desc)
        desc_cell.font = Font(size=9, italic=True, color='666666')
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
        desc_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Add sample placeholder rows
    placeholders = [
        [
            "Enter your question or prompt here...",
            "Enter the expected/reference answer here...",
            "Enter the AI-generated response here...",
            "Optional: Enter relevant context...",
            "Y"
        ],
        [
            "What is machine learning?",
            "Machine learning is a subset of artificial intelligence...",
            "Machine learning allows computers to learn from data...",
            "Computer science and AI fundamentals",
            "Y"
        ]
    ]
    
    for row_idx, placeholder_row in enumerate(placeholders, 6):
        for col_idx, placeholder in enumerate(placeholder_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=placeholder)
            if row_idx == 6:  # First placeholder row
                cell.font = Font(italic=True, color='999999')
            else:  # Example row
                cell.font = Font(color='2F5496')
    
    # Set column widths
    column_widths = [40, 60, 60, 40, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Set row heights
    ws.row_dimensions[5].height = 40
    
    # Add borders to header section
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(4, 8):
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border

def create_examples_sheet(wb):
    """Create examples sheet with sample data"""
    ws = wb.create_sheet("💡 Examples")
    
    # Title
    ws['A1'] = "Sample Test Cases - Examples"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:E1')
    
    # Create sample data
    sample_data = {
        'Prompt': [
            'What is photosynthesis?',
            'Explain how vaccines work',
            'What are the main causes of climate change?',
            'How does blockchain technology work?',
            'What is the difference between AI and machine learning?',
            'Describe the water cycle',
            'What are renewable energy sources?',
            'How do antibiotics work?',
            'What is quantum computing?',
            'Explain the greenhouse effect'
        ],
        'Reference_Response': [
            'Photosynthesis is the process by which plants convert sunlight, carbon dioxide, and water into glucose and oxygen using chlorophyll.',
            'Vaccines work by training the immune system to recognize and fight specific diseases by introducing a weakened or inactive form of the pathogen.',
            'Climate change is primarily caused by human activities including burning fossil fuels, deforestation, industrial processes, and agriculture.',
            'Blockchain is a distributed ledger technology that maintains a continuously growing list of records, secured by cryptography and consensus mechanisms.',
            'AI is the broader concept of machines simulating human intelligence, while machine learning is a subset of AI focused on learning from data.',
            'The water cycle involves evaporation, condensation, precipitation, and collection, continuously moving water through Earth\'s systems.',
            'Renewable energy sources include solar, wind, hydroelectric, geothermal, and biomass energy that naturally replenish over time.',
            'Antibiotics work by killing bacteria or inhibiting their growth through various mechanisms like disrupting cell walls or protein synthesis.',
            'Quantum computing uses quantum mechanical phenomena like superposition and entanglement to process information in fundamentally new ways.',
            'The greenhouse effect occurs when greenhouse gases trap heat in Earth\'s atmosphere, warming the planet\'s surface.'
        ],
        'Generated_Response': [
            'Plants use photosynthesis to make food by combining sunlight, CO2, and water to create sugar and release oxygen.',
            'Vaccines help the body build immunity by exposing it to safe versions of viruses or bacteria.',
            'Climate change is mainly caused by burning fossil fuels, which releases greenhouse gases into the atmosphere.',
            'Blockchain creates secure digital records that are distributed across multiple computers and cannot be easily changed.',
            'Artificial intelligence is the general field of making computers smart, while machine learning is specifically about learning from data.',
            'Water evaporates from oceans, forms clouds, falls as rain, and flows back to the ocean in a continuous cycle.',
            'Renewable energy comes from sources like the sun, wind, and water that don\'t run out.',
            'Antibiotics are medicines that kill harmful bacteria in the body to treat infections.',
            'Quantum computers use quantum physics principles to solve certain problems much faster than regular computers.',
            'The greenhouse effect happens when gases in the atmosphere trap heat from the sun, making Earth warmer.'
        ],
        'Context': [
            'Biology and plant science education',
            'Medical science and public health',
            'Environmental science and global issues',
            'Technology and cybersecurity',
            'Computer science and artificial intelligence',
            'Earth science and environmental studies',
            'Environmental science and sustainability',
            'Medical science and pharmacology',
            'Advanced computing and quantum physics',
            'Environmental science and climate studies'
        ],
        'Enable_Flag': ['Y'] * 10
    }
    
    # Convert to DataFrame and add to sheet
    df = pd.DataFrame(sample_data)
    
    # Add headers
    headers = ['Prompt', 'Reference_Response', 'Generated_Response', 'Context', 'Enable_Flag']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Set column widths
    column_widths = [35, 70, 70, 35, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Set row heights for better readability
    for row in range(4, 14):
        ws.row_dimensions[row].height = 60

def create_column_definitions_sheet(wb):
    """Create detailed column definitions sheet"""
    ws = wb.create_sheet("📖 Column Definitions")
    
    # Title
    ws['A1'] = "Column Definitions and Requirements"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:E1')
    
    # Column definitions data
    definitions = [
        {
            'Column': 'Prompt*',
            'Type': 'Required',
            'Description': 'The original question, instruction, or task given to the AI system',
            'Examples': 'What is machine learning?\nExplain photosynthesis\nWrite a summary of...',
            'Notes': 'Should be clear and specific. Used by Answer Relevance metric.'
        },
        {
            'Column': 'Reference_Response*',
            'Type': 'Required', 
            'Description': 'The expected, correct, or high-quality reference answer',
            'Examples': 'Expert-written answers\nHuman-validated responses\nGold standard outputs',
            'Notes': 'Used by BLEU, ROUGE, METEOR, BERT Score, and Accuracy metrics. Should be high-quality.'
        },
        {
            'Column': 'Generated_Response*',
            'Type': 'Required',
            'Description': 'The AI-generated response that you want to evaluate',
            'Examples': 'ChatGPT outputs\nCustom model responses\nAny AI-generated text',
            'Notes': 'Used by all metrics. This is what gets evaluated.'
        },
        {
            'Column': 'Context',
            'Type': 'Optional',
            'Description': 'Additional background information or context for the prompt',
            'Examples': 'Subject area\nDifficulty level\nSpecific domain context',
            'Notes': 'Improves Answer Relevance scoring. Can be left empty.'
        },
        {
            'Column': 'Enable_Flag',
            'Type': 'Optional',
            'Description': 'Y/N flag to enable or disable evaluation for this test case',
            'Examples': 'Y (enable)\nN (disable)',
            'Notes': 'Defaults to Y if not provided. Config file settings take precedence.'
        }
    ]
    
    # Headers
    headers = ['Column Name', 'Required?', 'Description', 'Examples', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Add definitions
    for row_idx, definition in enumerate(definitions, 4):
        ws.cell(row=row_idx, column=1, value=definition['Column']).font = Font(bold=True)
        
        type_cell = ws.cell(row=row_idx, column=2, value=definition['Type'])
        if definition['Type'] == 'Required':
            type_cell.font = Font(color='FF0000', bold=True)  # Red for required
        else:
            type_cell.font = Font(color='008000')  # Green for optional
        
        ws.cell(row=row_idx, column=3, value=definition['Description'])
        ws.cell(row=row_idx, column=4, value=definition['Examples'])
        ws.cell(row=row_idx, column=5, value=definition['Notes'])
        
        # Set alignment and wrapping
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Set column widths
    column_widths = [20, 15, 40, 30, 35]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Set row heights
    for row in range(4, 9):
        ws.row_dimensions[row].height = 80
    
    # Add additional information
    ws['A10'] = "IMPORTANT NOTES:"
    ws['A10'].font = Font(bold=True, size=12, color='2F5496')
    
    notes = [
        "",
        "• All text should be in plain text format (no special formatting)",
        "• Avoid very long responses (>5000 characters) for better performance", 
        "• Include diverse test cases for comprehensive evaluation",
        "• Reference responses should be high-quality and factually correct",
        "• Context field helps improve Answer Relevance but is not mandatory",
        "• Enable_Flag allows you to selectively run tests",
        "• Save file as .xlsx format before running evaluation",
        "",
        "METRICS DEPENDENCY SUMMARY:",
        "• BLEU, ROUGE, METEOR, BERT Score, Accuracy: Need Reference + Generated",
        "• Answer Relevance: Needs Prompt + Generated (+ optional Context)", 
        "• Toxicity: Only needs Generated response",
        "",
        "For detailed documentation, see README.md"
    ]
    
    for i, note in enumerate(notes, 11):
        ws[f'A{i}'] = note
        if note.startswith('•'):
            ws[f'A{i}'].font = Font(color='404040')
        elif note.endswith(':'):
            ws[f'A{i}'].font = Font(bold=True, color='2F5496')

if __name__ == "__main__":
    # Create the template
    template_path = create_excel_template()
    
    print("\n" + "=" * 50)
    print("EXCEL TEMPLATE CREATED SUCCESSFULLY!")
    print("=" * 50)
    print(f"📄 File: {template_path}")
    print("📋 Sheets included:")
    print("   • Instructions - How to use the template")
    print("   • Input Data - Main template for your data")
    print("   • Examples - Sample test cases")
    print("   • Column Definitions - Detailed field descriptions")
    print("\n💡 Next steps:")
    print("   1. Open the Excel file")
    print("   2. Read the Instructions sheet")
    print("   3. Fill in your data in the 'Input Data' sheet")
    print("   4. Save and run evaluation with:")
    print("      python src/main.py --input Excel_Input_Template.xlsx")
    print("=" * 50)