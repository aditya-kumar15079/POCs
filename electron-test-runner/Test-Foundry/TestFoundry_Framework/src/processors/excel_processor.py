"""
Excel Document Processor for TestFoundry Framework
Handles Excel spreadsheet parsing with data extraction and analysis
"""

import pandas as pd
from pathlib import Path
from typing import Dict, Any, List, Optional
import logging
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelProcessor:
    """Excel document processor with comprehensive data extraction"""
    
    def __init__(self, config: Dict[str, Any], logger: logging.Logger):
        """Initialize Excel processor
        
        Args:
            config: Configuration dictionary
            logger: Logger instance
        """
        self.config = config
        self.logger = logger
        self.excel_config = config.get('document_processing', {}).get('excel', {})
        
        # Settings
        self.read_all_sheets = self.excel_config.get('read_all_sheets', True)
        self.include_formulas = self.excel_config.get('include_formulas', False)
    
    async def process_document(self, file_path: Path) -> Dict[str, Any]:
        """Process Excel document and extract content
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary containing extracted content
        """
        try:
            self.logger.info(f"Processing Excel document: {file_path.name}")
            
            # Read Excel file
            if self.read_all_sheets:
                # Read all sheets
                excel_data = pd.read_excel(str(file_path), sheet_name=None, engine='openpyxl')
            else:
                # Read only first sheet
                excel_data = {0: pd.read_excel(str(file_path), engine='openpyxl')}
            
            # Process each sheet
            sheets = []
            for sheet_name, df in excel_data.items():
                sheet_content = await self._process_sheet(df, sheet_name, file_path)
                sheets.append(sheet_content)
            
            # Create document structure
            doc_content = {
                'name': file_path.stem,
                'file_path': str(file_path),
                'file_type': 'excel',
                'processor': 'ExcelProcessor',
                'sheets': sheets
            }
            
            # Calculate statistics
            total_cells = sum(sheet.get('total_cells', 0) for sheet in sheets)
            total_text = sum(len(sheet.get('content', '')) for sheet in sheets)
            
            doc_content['statistics'] = {
                'total_sheets': len(sheets),
                'total_cells': total_cells,
                'total_characters': total_text,
                'total_words': len(' '.join(sheet.get('content', '') for sheet in sheets).split()),
                'non_empty_cells': sum(sheet.get('non_empty_cells', 0) for sheet in sheets)
            }
            
            self.logger.info(f"Excel processing complete: {doc_content['statistics']['total_sheets']} sheets, "
                           f"{doc_content['statistics']['total_cells']} cells")
            
            return doc_content
            
        except Exception as e:
            self.logger.error(f"Error processing Excel document {file_path.name}: {e}")
            return {
                'name': file_path.stem,
                'file_path': str(file_path),
                'file_type': 'excel',
                'sheets': [],
                'error': str(e)
            }
    
    async def _process_sheet(self, df: pd.DataFrame, sheet_name: str, file_path: Path) -> Dict[str, Any]:
        """Process individual Excel sheet
        
        Args:
            df: Pandas DataFrame containing sheet data
            sheet_name: Name of the sheet
            file_path: Path to Excel file
            
        Returns:
            Dictionary containing sheet content
        """
        try:
            # Clean DataFrame
            df_clean = df.dropna(how='all').dropna(axis=1, how='all')
            
            # Convert to text representation
            content_parts = []
            
            # Add sheet name as header
            content_parts.append(f"Sheet: {sheet_name}")
            content_parts.append("=" * 50)
            
            # Add column headers
            if not df_clean.empty:
                headers = [str(col) for col in df_clean.columns]
                content_parts.append("Columns: " + " | ".join(headers))
                content_parts.append("-" * 50)
                
                # Add data rows (limit to reasonable number for text representation)
                max_rows = 100  # Limit rows for readability
                for idx, row in df_clean.head(max_rows).iterrows():
                    row_text = " | ".join(str(val) if pd.notna(val) else "" for val in row)
                    content_parts.append(f"Row {idx + 1}: {row_text}")
                
                # Add summary if there are more rows
                if len(df_clean) > max_rows:
                    content_parts.append(f"... and {len(df_clean) - max_rows} more rows")
            
            # Statistical analysis
            statistics = await self._analyze_sheet_data(df_clean)
            
            # Add statistics to content
            content_parts.append("\nData Summary:")
            content_parts.append(f"- Total Rows: {len(df_clean)}")
            content_parts.append(f"- Total Columns: {len(df_clean.columns)}")
            content_parts.append(f"- Non-empty Cells: {df_clean.count().sum()}")
            
            # Add numeric column statistics
            numeric_cols = df_clean.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                content_parts.append("\nNumeric Column Statistics:")
                for col in numeric_cols:
                    col_stats = df_clean[col].describe()
                    content_parts.append(f"- {col}: Mean={col_stats['mean']:.2f}, "
                                       f"Min={col_stats['min']:.2f}, Max={col_stats['max']:.2f}")
            
            # Create sheet content
            sheet_content = {
                'sheet_name': str(sheet_name),
                'content': "\n".join(content_parts),
                'data_frame': df_clean.to_dict('records'),  # Store structured data
                'columns': list(df_clean.columns),
                'total_rows': len(df_clean),
                'total_columns': len(df_clean.columns),
                'total_cells': len(df_clean) * len(df_clean.columns),
                'non_empty_cells': df_clean.count().sum(),
                'statistics': statistics,
                'data_types': {col: str(dtype) for col, dtype in df_clean.dtypes.items()}
            }
            
            # Extract formulas if enabled
            if self.include_formulas:
                formulas = await self._extract_formulas(file_path, sheet_name)
                sheet_content['formulas'] = formulas
                if formulas:
                    content_parts.append("\nFormulas:")
                    for formula in formulas:
                        content_parts.append(f"- {formula['cell']}: {formula['formula']}")
                    sheet_content['content'] = "\n".join(content_parts)
            
            return sheet_content
            
        except Exception as e:
            self.logger.error(f"Error processing sheet {sheet_name}: {e}")
            return {
                'sheet_name': str(sheet_name),
                'content': f"Error processing sheet: {str(e)}",
                'error': str(e)
            }
    
    async def _analyze_sheet_data(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Analyze sheet data and provide insights
        
        Args:
            df: Pandas DataFrame
            
        Returns:
            Dictionary containing analysis results
        """
        analysis = {
            'data_types': {},
            'missing_data': {},
            'unique_values': {},
            'numeric_summary': {},
            'text_summary': {}
        }
        
        try:
            for col in df.columns:
                col_data = df[col]
                
                # Data type analysis
                analysis['data_types'][col] = {
                    'dtype': str(col_data.dtype),
                    'is_numeric': pd.api.types.is_numeric_dtype(col_data),
                    'is_datetime': pd.api.types.is_datetime64_any_dtype(col_data),
                    'is_text': pd.api.types.is_string_dtype(col_data)
                }
                
                # Missing data analysis
                analysis['missing_data'][col] = {
                    'missing_count': col_data.isna().sum(),
                    'missing_percentage': (col_data.isna().sum() / len(col_data)) * 100
                }
                
                # Unique values analysis
                analysis['unique_values'][col] = {
                    'unique_count': col_data.nunique(),
                    'unique_percentage': (col_data.nunique() / len(col_data)) * 100
                }
                
                # Numeric analysis
                if pd.api.types.is_numeric_dtype(col_data):
                    try:
                        analysis['numeric_summary'][col] = {
                            'mean': float(col_data.mean()),
                            'median': float(col_data.median()),
                            'std': float(col_data.std()),
                            'min': float(col_data.min()),
                            'max': float(col_data.max()),
                            'quartiles': {
                                'q25': float(col_data.quantile(0.25)),
                                'q75': float(col_data.quantile(0.75))
                            }
                        }
                    except Exception:
                        pass
                
                # Text analysis
                elif pd.api.types.is_string_dtype(col_data):
                    try:
                        text_lengths = col_data.dropna().astype(str).str.len()
                        analysis['text_summary'][col] = {
                            'avg_length': float(text_lengths.mean()),
                            'max_length': int(text_lengths.max()),
                            'min_length': int(text_lengths.min()),
                            'total_chars': int(text_lengths.sum())
                        }
                    except Exception:
                        pass
        
        except Exception as e:
            self.logger.warning(f"Error in data analysis: {e}")
        
        return analysis
    
    async def _extract_formulas(self, file_path: Path, sheet_name: str) -> List[Dict[str, str]]:
        """Extract formulas from Excel sheet
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of the sheet
            
        Returns:
            List of formula dictionaries
        """
        formulas = []
        
        try:
            # Use openpyxl to access formulas
            workbook = openpyxl.load_workbook(str(file_path), data_only=False)
            
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas.append({
                                'cell': cell.coordinate,
                                'formula': cell.value,
                                'row': cell.row,
                                'column': cell.column
                            })
            
            workbook.close()
            
        except Exception as e:
            self.logger.warning(f"Error extracting formulas from {sheet_name}: {e}")
        
        return formulas
    
    def get_supported_extensions(self) -> List[str]:
        """Get list of supported file extensions
        
        Returns:
            List of supported extensions
        """
        return ['.xlsx', '.xls']