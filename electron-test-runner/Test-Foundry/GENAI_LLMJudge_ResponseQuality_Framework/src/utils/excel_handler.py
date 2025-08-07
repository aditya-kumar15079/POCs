"""
Excel handler for LLM Judge Framework.
Handles reading input Excel files and generating output reports.
"""

import os
import pandas as pd
from datetime import datetime
from typing import Dict, List, Tuple, Any, Optional

from src.utils.logger import Logger


class ExcelHandler:
    """Handler for Excel file operations."""

    def __init__(self, input_path: str, output_folder: str, scoring_config: Dict[str, Dict[str, Any]]):
        """
        Initialize the ExcelHandler.
        
        Args:
            input_path: Path to the input Excel file
            output_folder: Path to the output folder
            scoring_config: Configuration for scoring categories
        """
        self.input_path = input_path
        self.output_folder = output_folder
        self.scoring_config = scoring_config
        
        # Create output folder if it doesn't exist
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

    def read_test_cases(self) -> Tuple[pd.DataFrame, Optional[Dict[str, bool]]]:
        """
        Read test cases from the input Excel file.
        
        Returns:
            Tuple containing:
                - DataFrame with test cases
                - Optional Dict with metric configurations (if present in Excel)
        
        Raises:
            FileNotFoundError: If the input file doesn't exist
            ValueError: If the input file has invalid format
        """
        try:
            # Check if file exists
            if not os.path.exists(self.input_path):
                raise FileNotFoundError(f"Input file not found at {self.input_path}")
            
            # Read Excel file
            test_cases_df = pd.read_excel(self.input_path, sheet_name=0)  # Read first sheet regardless of name
            
            # Validate required columns
            required_cols = ["TestID", "Prompt", "ReferenceAnswer", "GeneratedAnswer"]
            missing_cols = [col for col in required_cols if col not in test_cases_df.columns]
            if missing_cols:
                raise ValueError(f"Input file missing required columns: {', '.join(missing_cols)}")
            
            # Try to read metrics configuration if exists
            metrics_config = None
            try:
                metrics_df = pd.read_excel(self.input_path, sheet_name="Metrics")
                metrics_config = {}
                for _, row in metrics_df.iterrows():
                    if "Metric" in row and "Enabled" in row:
                        # Convert to boolean
                        metrics_config[row["Metric"]] = (
                            row["Enabled"] if isinstance(row["Enabled"], bool) 
                            else str(row["Enabled"]).lower() in ["true", "yes", "1"]
                        )
            except Exception as e:
                Logger.info(f"No metrics configuration found in Excel: {str(e)}")
            
            return test_cases_df, metrics_config
        
        except Exception as e:
            Logger.error(f"Error reading Excel file: {str(e)}")
            raise
    
    def generate_output_report(self, results_df: pd.DataFrame, file_format: str = "LLM_Judge_ResponseQuality_Results_{timestamp}.xlsx") -> str:
        """
        Generate color-coded output report.
        
        Args:
            results_df: DataFrame containing evaluation results
            file_format: Format string for output file name
        
        Returns:
            Path to the generated output file
        """
        try:
            # Create output filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = file_format.format(timestamp=timestamp)
            output_path = os.path.join(self.output_folder, output_filename)
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Create a summary dataframe first
                summary_df = self._create_summary_df(results_df)
                
                # Create metrics explanation dataframe
                metrics_explanation_df = self._create_metrics_explanation_df()
                
                # Write summary to Excel
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                
                # Write metrics explanation to Excel
                metrics_explanation_df.to_excel(writer, sheet_name="Metrics Explanation", index=False)
                
                # Write detailed results to Excel
                results_df.to_excel(writer, sheet_name="Detailed Results", index=False)
                
                # Access workbook and worksheets
                workbook = writer.book
                summary_sheet = writer.sheets["Summary"]
                metrics_sheet = writer.sheets["Metrics Explanation"]
                detail_sheet = writer.sheets["Detailed Results"]
                
                # Format column widths in summary sheet
                self._format_worksheet(summary_sheet, summary_df)
                
                # Format column widths in metrics explanation sheet
                self._format_worksheet(metrics_sheet, metrics_explanation_df)
                
                # Format column widths in detail sheet
                self._format_worksheet(detail_sheet, results_df)
                
                # Apply cell formatting and coloring to the sheets
                self._apply_cell_formatting(detail_sheet, results_df)
                self._apply_color_coding_to_summary(summary_sheet, summary_df)
                
                # Format metrics explanation sheet
                self._format_metrics_explanation_sheet(metrics_sheet, metrics_explanation_df, workbook)
                
                # Add title and formatting to summary sheet
                title_font = self._get_title_font(workbook)
                summary_sheet['A1'].font = title_font
                
                # Add explanation at the bottom of summary sheet
                row_idx = len(summary_df) + 3
                summary_sheet.cell(row=row_idx, column=1).value = "Scoring Categories:"
                summary_sheet.cell(row=row_idx, column=1).font = self._get_bold_font(workbook)
                
                # Add scoring explanations
                for i, (category, config) in enumerate(self.scoring_config.items()):
                    row_idx += 1
                    cell = summary_sheet.cell(row=row_idx, column=1)
                    cell.value = f"{config['min']}-{config['max']}: {category.capitalize()}"
                
                # Add explanation for gray cells (skipped evaluations)
                row_idx += 2
                cell = summary_sheet.cell(row=row_idx, column=1)
                cell.value = "Note: Gray cells indicate metrics that could not be evaluated (typically due to missing reference answers)"
                cell.font = self._get_bold_font(workbook)
                
            Logger.info(f"Generated output report at {output_path}")
            return output_path
        
        except Exception as e:
            Logger.error(f"Error generating output report: {str(e)}")
            raise
    
    def _create_summary_df(self, results_df: pd.DataFrame) -> pd.DataFrame:
        """
        Create a summary DataFrame with test results.
        
        Args:
            results_df: DataFrame with detailed results
            
        Returns:
            Summary DataFrame
        """
        # Get metric columns (they end with "Score")
        metric_cols = [col for col in results_df.columns if col.endswith("Score")]
        
        # Initialize summary data
        summary_data = []
        
        # Add average scores for each metric - only including non-null values
        avg_row = {"Metric": "Average Score"}
        for col in metric_cols:
            metric_name = col.replace("Score", "")
            # Calculate average only for non-null values
            valid_scores = results_df[col].dropna()
            if len(valid_scores) > 0:
                avg_row[metric_name] = valid_scores.mean()
            else:
                avg_row[metric_name] = None  # No valid scores
        summary_data.append(avg_row)
        
        # Add test case specific scores
        for idx, row in results_df.iterrows():
            test_row = {"Metric": f"Test Case: {row['TestID']}"}
            for col in metric_cols:
                metric_name = col.replace("Score", "")
                test_row[metric_name] = row[col]
            summary_data.append(test_row)
        
        # Create DataFrame
        summary_df = pd.DataFrame(summary_data)
        
        return summary_df
    
    def _create_metrics_explanation_df(self) -> pd.DataFrame:
        """
        Create a DataFrame with metric explanations.
        
        Returns:
            DataFrame with metric explanations
        """
        # Define metrics explanations
        metrics_data = [
            {
                "Metric": "Accuracy",
                "Description": "Accuracy assesses the factual correctness of the generated text compared to the reference answer.",
                "Thresholds": "Green: ≥90%\nYellow: ≥75%\nAmber: ≥50%\nRed: <50%",
                "Interpretation": "Higher scores are better. 0-100 scale.\n• 90-100 = Excellent accuracy\n• 75-89 = Good accuracy\n• 50-74 = Moderate accuracy\n• 0-49 = Poor accuracy"
            },
            {
                "Metric": "Coherence",
                "Description": "Coherence evaluates the logical flow, consistency, and organization of the generated text, assessing if it's well-structured and follows a clear progression.",
                "Thresholds": "Green: ≥90%\nYellow: ≥75%\nAmber: ≥50%\nRed: <50%",
                "Interpretation": "Higher scores are better. 0-100 scale.\n• 90-100 = Excellent logical flow\n• 75-89 = Good coherence\n• 50-74 = Moderate coherence\n• 0-49 = Poor coherence"
            },
            {
                "Metric": "Relevance",
                "Description": "Relevance measures how well the generated text addresses the input prompt, using semantic similarity to assess topical alignment.",
                "Thresholds": "Green: ≥90%\nYellow: ≥75%\nAmber: ≥50%\nRed: <50%", 
                "Interpretation": "Higher scores are better. 0-100 scale.\n• 90-100 = Highly relevant to prompt\n• 75-89 = Good relevance\n• 50-74 = Moderately relevant\n• 0-49 = Poorly relevant"
            },
            {
                "Metric": "Faithfulness",
                "Description": "Faithfulness measures the absence of hallucinations or made-up information, ensuring the response doesn't contain facts not present in the context or reference.",
                "Thresholds": "Green: ≥90%\nYellow: ≥75%\nAmber: ≥50%\nRed: <50%",
                "Interpretation": "Higher scores are better. 0-100 scale.\n• 90-100 = No hallucinations\n• 75-89 = Minor inaccuracies\n• 50-74 = Some hallucinations\n• 0-49 = Major hallucinations"
            },
            {
                "Metric": "Bias",
                "Description": "Bias evaluates whether the generated text exhibits unfair preference or prejudice toward particular viewpoints, groups, or perspectives.",
                "Thresholds": "Green: ≥90%\nYellow: ≥75%\nAmber: ≥50%\nRed: <50%",
                "Interpretation": "Higher scores are better. 0-100 scale.\n• 90-100 = Completely unbiased\n• 75-89 = Minimal bias\n• 50-74 = Noticeable bias\n• 0-49 = Significant bias"
            },
            {
                "Metric": "Toxicity",
                "Description": "Toxicity assesses whether the generated text contains harmful, offensive, or inappropriate content including hate speech, profanity, or content promoting illegal activities.",
                "Thresholds": "Green: ≥90%\nYellow: ≥75%\nAmber: ≥50%\nRed: <50%",
                "Interpretation": "Higher scores are better. 0-100 scale.\n• 90-100 = Non-toxic content\n• 75-89 = Minimal concerns\n• 50-74 = Some problematic content\n• 0-49 = Significant toxicity"
            }
        ]
        
        # Create DataFrame
        metrics_df = pd.DataFrame(metrics_data)
        
        return metrics_df
    
    def _format_worksheet(self, worksheet, df: pd.DataFrame):
        """
        Format worksheet column widths.
        
        Args:
            worksheet: Worksheet to format
            df: DataFrame with data
        """
        # For each column in the dataframe
        for idx, column in enumerate(df.columns):
            # Find the maximum length of any cell in this column
            max_length = max(
                df[column].astype(str).map(len).max(),  # Max length of data
                len(str(column))  # Length of column name
            ) + 3  # Add a little extra space
            
            # Set column width
            col_letter = worksheet.cell(row=1, column=idx+1).column_letter
            if max_length > 100:  # Cap at 100 for very long content
                worksheet.column_dimensions[col_letter].width = 100
            else:
                worksheet.column_dimensions[col_letter].width = max_length
        
        # Set row height for better readability
        for row in worksheet.rows:
            worksheet.row_dimensions[row[0].row].height = 20
    
    def _apply_cell_formatting(self, worksheet, df: pd.DataFrame):
        """
        Apply cell formatting and coloring to detailed results.
        
        Args:
            worksheet: Worksheet to format
            df: DataFrame with data
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Create header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Apply header styling
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border
        
        # Apply data cell styling and colors
        for row_idx in range(2, len(df) + 2):
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
                
                # Apply color to score cells
                if str(col_name).endswith("Score"):
                    try:
                        value = cell.value
                        if value is None or (isinstance(value, float) and pd.isna(value)):
                            # Apply gray for skipped evaluations
                            cell.fill = self._get_score_fills()["gray"]
                        else:
                            score = float(value)
                            # Apply color based on score
                            self._apply_score_color(cell, score)
                    except (ValueError, TypeError):
                        # Not a valid number
                        cell.fill = self._get_score_fills()["gray"]
    
    def _apply_color_coding_to_summary(self, worksheet, df: pd.DataFrame):
        """
        Apply color coding to summary sheet, including the tabular section.
        
        Args:
            worksheet: Worksheet to format
            df: DataFrame with data
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Create header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Apply header styling
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border
        
        # Apply data cell styling and colors
        for row_idx in range(2, len(df) + 2):
            # Apply styling to Metric column (first column)
            cell = worksheet.cell(row=row_idx, column=1)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            cell.font = Font(bold=True)
            
            # Apply color coding to metric score cells (second column onwards)
            for col_idx in range(2, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                
                # Get score value
                try:
                    value = cell.value
                    if value is None or (isinstance(value, float) and pd.isna(value)):
                        # Apply gray for skipped evaluations
                        cell.fill = self._get_score_fills()["gray"]
                    else:
                        score = float(value)
                        # Apply color based on score
                        self._apply_score_color(cell, score)
                except (ValueError, TypeError):
                    # Not a valid number
                    cell.fill = self._get_score_fills()["gray"]
    
    def _apply_score_color(self, cell, score):
        """
        Apply color to a cell based on score value.
        
        Args:
            cell: Cell to format
            score: Score value
        """
        # Get fills for different score categories
        fills = self._get_score_fills()
        
        # Determine which category the score falls into
        for category, config in self.scoring_config.items():
            min_val = config["min"]
            max_val = config["max"]
            
            if min_val <= score <= max_val:
                cell.fill = fills[config["color"]]
                break
    
    def _get_score_fills(self):
        """
        Get color fills for different score categories.
        
        Returns:
            Dict of colors to PatternFill objects
        """
        from openpyxl.styles import PatternFill
        
        return {
            "green": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            "light_green": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "amber": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "red": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "gray": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # For skipped/unavailable scores
        }
    
    def _format_metrics_explanation_sheet(self, worksheet, df: pd.DataFrame, workbook):
        """
        Format the metrics explanation sheet.
        
        Args:
            worksheet: Worksheet to format
            df: DataFrame with data
            workbook: Excel workbook
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Add title to the sheet
        worksheet.merge_cells('A1:D1')
        title_cell = worksheet['A1']
        title_cell.value = "GenAI Testing Framework Metrics Explanation"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Shift all data down one row to make room for the title
        for row in range(df.shape[0] + 1, 1, -1):
            for col in range(1, df.shape[1] + 1):
                worksheet.cell(row=row+1, column=col).value = worksheet.cell(row=row, column=col).value
        
        # Create header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Apply header styling
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=2, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border
        
        # Apply data cell styling
        for row_idx in range(3, len(df) + 3):
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
                
                # Make metric names bold
                if col_idx == 1:  # Metric column
                    cell.font = Font(bold=True)
        
        # Adjust row heights for better readability
        worksheet.row_dimensions[1].height = 30  # Title row
        for row in range(2, len(df) + 3):
            worksheet.row_dimensions[row].height = 75  # Data rows
        
        # Add a note about gray cells at the bottom
        note_row = len(df) + 4
        worksheet.merge_cells(f'A{note_row}:D{note_row}')
        note_cell = worksheet.cell(row=note_row, column=1)
        note_cell.value = "Note: Some metrics (Accuracy, Faithfulness) require a reference answer and will be skipped when none is provided."
        note_cell.font = Font(italic=True)
        note_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    def _get_title_font(self, workbook):
        """Get font for title cells."""
        return workbook.create_font(name='Calibri', size=14, bold=True, color='4472C4')
    
    def _get_bold_font(self, workbook):
        """Get bold font for emphasis."""
        return workbook.create_font(name='Calibri', size=11, bold=True)